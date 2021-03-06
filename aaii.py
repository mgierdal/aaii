import os

print(os.curdir)

os.chdir('.')

import urllib2
import sys
import requests
from secrets import ID, PASSWORD
import shutil
import re
import bs4
import csv
import pandas as pd
import openpyxl
import xlrd
import collections


STOCK_IDEAS_URL = r'http://www.aaii.com/stockideas' #+ '?a=menubarHome'

# all screens
PERFORMANCE_ROOT    = r'http://www.aaii.com/stockideas/performance' # also redirected from 
PERFORMANCE_MONTHLY = r'http://www.aaii.com/files/spreadsheets/stockideas/monthlyperformance.xlsx'
PERFORMANCE_ANNUALLY   = r'http://www.aaii.com/files/spreadsheets/stockideas/annualperformance.xlsx'

AAII_BASE_URL = r'http://www.aaii.com'
SCREEN_MAIN_URL = r'http://www.aaii.com/stock-screens'
screens_mmenubarhome_url = r'http://www.aaii.com/stock-screens?a=menubarHome'
PERFORMANCE_HISTORY_URL = r'http://www.aaii.com/stock-screens/performance'

# search for http://www.aaii.com/stock-screens/screendata/CANSLIMRev

""" NEW URL PATTERN
http://www.aaii.com/stockideas/screendata/MAGNETComplex
http://www.aaii.com/files/spreadsheets/stockideas/passingcompanies/MAGNETComplex.xlsx
"""
# individual screens
r'http://www.aaii.com/stockideas/screendata/CashRichFirms'
r'http://www.aaii.com/files/spreadsheets/stockideas/passingcompanies/CashRichFirms.xlsx'

# Fill in your details here to be posted to the login form.
payload = {
    'inUserName': ID,
    'inUserPass': PASSWORD,
}

#url = LOGIN_URL


def download_and_save_file(url):
    """
    from https://stackoverflow.com/questions/16694907/how-to-download-large-file-in-python-with-requests-py
    """
    local_filename = url.split('/')[-1]
    r = requests.get(url, stream=True)
    with open(local_filename, 'wb') as f:
        shutil.copyfileobj(r.raw, f)

    return local_filename

def download_page(url):
    """
    download page
    """
    response = requests.get(STOCK_IDEAS_URL)
    return response.text.split('\r\n')


def download_page_(url):
    """
    download page
    """
    opener = urllib2.build_opener(urllib2.HTTPCookieProcessor())
    response = opener.open(url)
    page = response.read()
    return page

def save_page(html, filename):
    """
    save page
    """
    import types
    assert type(html) == types.ListType
    with open(filename, 'w') as f:
        f.writelines(html)
##################################
def find_xlsx_href(s):
    from re import findall
    ptrn = r'(href.+?xlsx)'
    return findall(ptrn, s)
def extract_screen_urls(s):
    from re import findall
    ptrn = r'(href.+?screendata.?")'
    return findall(ptrn, s)
def find_screendata_href(s):
    """
    <span class="screenname"><a href="/stock-screens/screendata/MAGNETSimple                  ">MAGNET 
    """
    ptrn = r'(<a\shref.+?\">)'
    ret = re.findall(ptrn, s)
    ret = [x for x in ret if re.search('screendata', x)]
    ret = [re.findall('".+?"',x)[0] for x in ret]
    ret = [x.strip('"').strip() for x in ret]
    return list(set(ret))
def extract_screen_title(html):
    # <h1 id="page_title" style="visibility: visible;">Dual Cash Flow Screen</h1>
    # <h1 id="page_title">O'Shaughnessy: Growth Market Leaders Screen</h1>
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    title = soup.find('h1', {'id':'page_title'})
    if title is None:
        title = soup.find('h1')
        if title is None:
            msg = '<h1 id="page_title" tag not found'
            raise ValueError(msg)
    return title.text
def extract_screen_origination_date(s):
    # <div align="right"><strong>Data as of 9/29/2017</strong></div>
    ptrn = r'<div.+?<strong>(Data as of [0-9/]+)</strong></div>'
    return re.findall(ptrn, s)
def extract_table(html):
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    table = soup.find('table')
    return [[col.text for col in row.findAll('td')] for row in table.findAll('tr')]
##################################
def get_aaii_screen_page():
    url = SCREEN_MAIN_URL
    #print url
    opener = urllib2.build_opener(urllib2.HTTPCookieProcessor())
    response = opener.open(url)
    page = response.read()
    return page
def save_aaii_screen_page_(html):
    save_page(html, 'stock_screens.html')
    
def make_screen_info(screen_page_url):
    """
    TODO refactor to separate IO from functional code
    input format: 'http://www.aaii.com/stockideas/screendata/ValueEstGrowth'
    """
    info = {}
    source_url = screen_page_url
    screen_label = screen_page_url.split(r'/')[-1]
    print 'XTERN: {}'.format((screen_label, source_url))
    info['label'] = screen_label
    info['url'] = source_url
    # download screen's webpage
    page = urllib2.build_opener(urllib2.HTTPCookieProcessor()).open(source_url).read()
    #if screen_label == 'MAGNETComplexRev': print page[:10]
    # find screen's name by finding webpage's title
    try:
        screen_title = extract_screen_title(page)
    except Exception, e:
        screen_title = ''
        print ('Error at {}: {}'. format(screen_label, str(e)))
        #print page[:1000]
        raise
    print 'TITLE: [{}]'.format(screen_title)
    info['full_name'] = screen_title
    date = re.sub('Data as of ','',extract_screen_origination_date(page)[0])
    info['origin_date'] = date
    #print len(page), screen_title, date
    page = get_aaii_screen_passing_companies_page(screen_label)
    #print len(page)
    # IO is here:
    save_aaii_screen_composition_page(screen_label, page)
    # http://www.aaii.com/stock-screens/screendata/DualCashFlow
    # http://www.aaii.com/stockscreens/passingco/DualCashFlow
    info['composition'] = extract_table(page)
    if info['composition'][-1]:
        info['composition'] = info['composition'][:-1]
    if info['composition'][0]==['No stocks passed the screen this month']:
        info['composition'] = None
    return info
def get_aaii_screen_passing_companies_page(label):
    # http://www.aaii.com/stockscreens/passingco/DualCashFlow
    url = r''.join([r'http://www.aaii.com/stockscreens/passingco/', label])
    #print url
    opener = urllib2.build_opener(urllib2.HTTPCookieProcessor())
    response = opener.open(url)
    page = response.read()
    return page
def save_aaii_screen_composition_page_(label, html):
    """
    original
    """
    file_name = label + r'.html'
    with open(file_name, 'w') as f:
        f.writelines(html)
def save_aaii_screen_composition_page(label, html):
    """ TO DO replac with direct save_page()"""
    file_name = label + r'.html'
    save_page(html, file_name)
    

##################################

DOWNLOAD = True

if __name__=='__main__':
    # start
    print 'DUPA'
    page_html = download_page(STOCK_IDEAS_URL)
    print len(page_html)
    root = bs4.BeautifulSoup(reduce(lambda x,y:x+y,page_html), 'lxml')
    
    save_page([line.encode('utf-8') for line in page_html], 'performance.html') # stock_screens.html
    # 2 performance Excel sheets
    print extract_screen_urls(reduce(lambda x,y:x+y, page_html))
    performance_sheets = [r''.join([AAII_BASE_URL, re.sub(r'href.*?"','',x)]) for x in find_xlsx_href(reduce(lambda x,y:x+y, page_html))]
    
    print 'Perf Sheets:', performance_sheets
    for sheet in performance_sheets[:]:
        #print sheet
        if DOWNLOAD:
            download_and_save_file(sheet)
            print '{} downloaded'.format(sheet)
    # individual screen files
    sys.exit()
    page_html = get_aaii_screen_page() # from SCREEN_MAIN_URL
    save_page(page_html, 'stock_screens.html')
    screen_webpages = [r''.join([AAII_BASE_URL, re.sub(r'href.*?"','',x)]) for x in find_screendata_href(page_html)[:]]
    screens = [make_screen_info(screen) for screen in screen_webpages[:]]
    for screen in screen_webpages[:]:
        #make_screen_info(screen)
        pass
    #sys.exit()
    # saving screens' passing companies
    for screen in screens[:]:
        print screen.keys()
        if screen['composition']:
            print screen['composition'][-1]
            file_name = screen['label'] + r'.csv'
            if DOWNLOAD:
                with open(file_name, 'wb') as fout:
                    wr = csv.writer(fout)
                    wr.writerows(screen['composition'])
    
##    Driehaus.html
##    Foolish8.html
##    Foolish8Rev.html
##    Rule1.html
    
    listing = os.listdir('.')
    df_aggr = pd.DataFrame(None)
    for fn in [x for x in listing if os.path.splitext(x)[-1] in ['.csv']][:]:
        df = pd.read_csv(os.path.abspath(fn))
        df['screen'] = os.path.splitext(fn)[0]
        df_aggr = df_aggr.append(df, ignore_index=True)
    print df_aggr.shape
    #counts = collections.Counter(df_aggr.Ticker.values)
    #print counts
    #print df_aggr.groupby('Ticker')['Ticker'].filter(lambda x:x.count()>1)
    sys.exit()
    ##################
    book = xlrd.open_workbook("annualperformance.xlsx")
    print("The number of worksheets is {0}".format(book.nsheets))
    print("Worksheet name(s): {0}".format(book.sheet_names()))
    sh = book.sheet_by_index(0)
    print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
    print("Cell D30 is {0}".format(sh.cell_value(rowx=29, colx=3)))
    for rx in range(sh.nrows):
        pass#print(sh.row(rx))
    sh.col(2)[4]

    #
    """
        import openpyxl

        wb = openpyxl.load_workbook('annualperformance.xlsx', read_only=True)
        print wb.get_sheet_names()
        ws = wb.get_active_sheet()
        print ws.title
        print

        def get_column(ws, col_idx):
            #here you iterate over the rows in the specific column
            ret =  [ [ws["{}{}".format(column, row)].value 
                      for column in col_idx] 
                    for row in range(1,ws.max_row)]
            return reduce(lambda x,y: x+y, ret)

        print get_column(ws, 'C')

        [(i,x) for i,x in enumerate(get_column(ws, 'C')) if x == 'YTD*']
    """
    sys.exit()
    
    # Use 'with' to ensure the session context is closed after use.
    with requests.Session() as s:
        stock_ideas = s.post(STOCK_IDEAS_URL, data=payload)
        # print the html returned or something more intelligent to see if it's a successful login page.
        #print (p.text)

        performance = s.post(PERFORMANCE_ROOT, data=payload)

        # An authorised request.
        #r = s.get(PERFORMANCE_MONTHLY)
        #print (r.text)
        
        download_and_save_file(PERFORMANCE_MONTHLY)
        download_and_save_file(PERFORMANCE_ANNUALLY)
